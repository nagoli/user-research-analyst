import re
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from typing import Dict, List, Tuple
from user_research_analyst.campaign.question_parsing import parse_questions


def load_results(results_file: str) -> Dict[str, Dict]:
    """Load results from a JSON file"""
    with open(results_file, 'r', encoding='utf-8') as f:
        return json.load(f)

def create_excel_report(questions: List[Tuple[str, str]], results_files: List[str], output_file: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Results"

    # Write headers
    ws.cell(row=1, column=1, value="File Name")
    ws.cell(row=1, column=2, value="Features")
    for col, (_, question_text) in enumerate(questions, start=3):
        ws.cell(row=1, column=col, value=question_text)

    # Define fill colors
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Process each results file
    for row, results_file in enumerate(results_files, start=2):
        file_name = os.path.splitext(os.path.basename(results_file))[0]
        ws.cell(row=row, column=1, value=file_name)

        results = load_results(results_file)
        for col, (question_id, _) in enumerate(questions, start=3):
            cell = ws.cell(row=row, column=col)
            if question_id in results:
                result = results[question_id]
                if result['found']:
                    cell.value = result['answer']
                    if result['confidence'] == 'medium':
                        cell.fill = orange_fill
                    elif result['confidence'] == 'low':
                        cell.fill = red_fill

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length, 100)  # Cap at 100 for readability

    wb.save(output_file)

def test_report_builder():
    # Example usage
    questions_file = 'questions.txt'
    results_directory = 'results'
    output_file = 'analysis_report.xlsx'

    questions = parse_questions(questions_file)
    results_files = [os.path.join(results_directory, f) for f in os.listdir(results_directory) if f.endswith('.json')]

    create_excel_report(questions, results_files, output_file)





if __name__ == "__main__":
    test_report_builder()




