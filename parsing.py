import re
import json
import os
from openpyxl import Workbook
from openpyxl.styles import PatternFill
from typing import Dict, List, Tuple

def parse_questions(file_path: str) -> List[Tuple[str, str]]:
    """
    Parse le fichier de questions et retourne une liste de tuples (id, question)
    
    Args:
        file_path: Chemin vers le fichier questions.txt
    
    Returns:
        List[Tuple[str, str]]: Liste de tuples (id, question)
    """
    questions = []
    current_id = ""
    current_question = []
    
    with open(file_path, 'r', encoding='utf-8') as f:
        for line in f:
            line = line.strip()
            
            # Skip empty lines
            if not line:
                continue
                
            # Check if line starts with question ID (e.g., "A10-", "B20-", etc.)
            id_match = re.match(r'^([A-Z]\d{2})-\s*(.*)', line)
            
            if id_match:
                # If we have a previous question stored, add it to the list
                if current_id and current_question:
                    questions.append((current_id, ' '.join(current_question)))
                
                # Start new question
                current_id = id_match.group(1)
                current_question = [id_match.group(2)]
            
            # If line starts with bullet point or is continuation of question
            elif line.startswith('•'):
                if current_question:
                    current_question.append(line.strip('•').strip())
            elif current_question:
                current_question.append(line)
    
    # Add the last question
    if current_id and current_question:
        questions.append((current_id, ' '.join(current_question)))
    
    return questions






def load_results(results_file: str) -> Dict[str, Dict]:
    """Load results from a JSON file"""
    with open(results_file, 'r', encoding='utf-8') as f:
        return json.load(f)

def create_excel_report(questions: List[Tuple[str, str]], results_files: List[str], output_file: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "Analysis Results"

    # Write headers
    ws.cell(row=1, column=1, value="File Name")
    ws.cell(row=1, column=2, value="Features")
    for col, (_, question_text) in enumerate(questions, start=3):
        ws.cell(row=1, column=col, value=question_text)

    # Define fill colors
    orange_fill = PatternFill(start_color="FFA500", end_color="FFA500", fill_type="solid")
    red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

    # Process each results file
    for row, results_file in enumerate(results_files, start=2):
        file_name = os.path.splitext(os.path.basename(results_file))[0]
        ws.cell(row=row, column=1, value=file_name)

        results = load_results(results_file)
        for col, (question_id, _) in enumerate(questions, start=3):
            cell = ws.cell(row=row, column=col)
            if question_id in results:
                result = results[question_id]
                if result['found']:
                    cell.value = result['answer']
                    if result['confidence'] == 'medium':
                        cell.fill = orange_fill
                    elif result['confidence'] == 'low':
                        cell.fill = red_fill

    # Adjust column widths
    for column_cells in ws.columns:
        length = max(len(str(cell.value)) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(length, 100)  # Cap at 100 for readability

    wb.save(output_file)

# Example usage
questions_file = 'questions.txt'
results_directory = 'results'
output_file = 'analysis_report.xlsx'

questions = parse_questions(questions_file)
results_files = [os.path.join(results_directory, f) for f in os.listdir(results_directory) if f.endswith('.json')]

create_excel_report(questions, results_files, output_file)








